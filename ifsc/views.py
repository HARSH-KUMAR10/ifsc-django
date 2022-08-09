from django.http import JsonResponse
import json
import openpyxl
from . import config


def ifsc_home(request,*args,**kwargs):
    return JsonResponse({"message":"Welcome to GDNA project by harsh kumar","statusCode":200})

def ifsc_find(request,*args,**kwargs):
    config.apiCount = config.apiCount+1
    try:
        response={}
        requestCode = json.loads(request.body)['code']
        if(len(requestCode)==0):
            return JsonResponse({"statusCode":404,"message":"Bank code not found, please enter bank code"})
        found = False
        if requestCode in config.cache:
            found = True
            response["bank"] = config.cache[requestCode]
        if(found):
            print("Available in cache.")
            config.ifscCount[requestCode] = config.ifscCount[requestCode]+1
            print('Cached data:',config.cache,'\n\nApi Hit Counts',config.apiCount,'\n\nIFSC Hit counts',config.ifscCount)
            return JsonResponse({"statusCode":200,"message":"Your bank details are.","data":response})
        else:
            config.ifscCount[requestCode] = 1
            print("Not available in cache, looking in db.")
            wb = openpyxl.load_workbook('C:/Users/admin/Desktop/Internship/WebKnot/django/GDNA/backend/ifsc/ifcb/IFCB2009_04.xlsx')
            print(wb.sheetnames)
            sheetName = wb.sheetnames[0]
            rowCount = wb[sheetName].max_row
            colCount = wb[sheetName].max_column
            vals=[]
            for i in range(1,rowCount+1):
                sheet = wb[sheetName]._get_cell(row=i,column=2)
                if(sheet.value==requestCode):
                    for j in range(1,colCount):
                        vals.append(wb[sheetName]._get_cell(row=i,column=j).value)
                    break
            if(len(vals)==0):
                return JsonResponse({"statusCode":404,"message":"Your bank details not found"})
            else:
                response["bank"] = vals
                config.cache[requestCode] = vals
                print('Cached data:',config.cache,'\n\nApi Hit Counts',config.apiCount,'\n\nIFSC Hit counts',config.ifscCount)
                return JsonResponse({"statusCode":200,"message":"Your bank details are.","data":response})
    except:
        return JsonResponse({"statusCode":500,"message":"unable to process, some error occured"})